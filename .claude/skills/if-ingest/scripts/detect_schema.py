"""从一份历史设计书（参考文档）样本探测表格结构，生成 projects/<name>/config.yaml 草稿。

要找的：映射表工作表、header_row、两个字段块（源+目标）、SAP 侧 label。

用法：
  python3 scripts/detect_schema.py <sample.xlsx> --project <name> [--out config.yaml]
"""
from __future__ import annotations

import argparse
import re
import subprocess
import sys
import tempfile
import unicodedata
import warnings
from pathlib import Path
from typing import Any

import openpyxl
import yaml

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ---- 语义同义词集 ----
FIELD_SEMANTICS_VOCAB: dict[str, list[str]] = {
    "no":     ["№", "No.", "No", "NO", "序号", "项次", "#"],
    "name":   ["項目名称", "項目名", "Field Name", "Column Name", "字段名称", "字段名", "名称"],
    "struct": ["構造", "Table", "表名", "テーブル", "构造", "Structure"],
    "tech":   ["技術名称", "技術名", "項目コード", "Technical Name", "Field Code", "技术名称", "技术名", "字段代码"],
    "length": ["文字数", "桁数", "バイト数", "Leng", "Length", "Len", "Byte", "长度", "バイト"],
    "type":   ["属性", "Type", "データ型", "类型", "型"],
}

# 映射表工作表名关键词
MAPPING_SHEET_KEYWORDS = ["項目マッピング", "マッピング", "Mapping", "mapping", "映射", "Map"]

# SAP 侧 label 正则
SAP_LABEL_PATTERNS = [r"部品\s*SAP", r"ＳＡＰ"]

# 辅助列同义词
AUX_VOCAB = {
    "conv_spec":          ["変換仕様"],
    "conv_current":       ["現行編集仕様"],
    "sap_digits":         ["桁数"],
    "sap_code_system":    ["コード体系"],
    "sap_supplement":     ["補足・その他", "補足"],
    "unrealizable_no":    ["No."],
    "unrealizable_class": ["分類"],
    "remark":             ["備考", "Remark", "备注"],
}


# ---- 基础工具 ----

def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _match_vocab(text: str | None, vocab: list[str]) -> bool:
    if not text:
        return False
    t = unicodedata.normalize("NFKC", str(text)).strip()
    return any(t == w or t.startswith(w) for w in vocab)


def _semantic_for(text: str | None) -> str | None:
    if not text:
        return None
    for sem, vocab in FIELD_SEMANTICS_VOCAB.items():
        if _match_vocab(text, vocab):
            return sem
    return None


def _ensure_xlsx(path: Path) -> Path:
    if path.suffix.lower() == ".xlsx":
        return path
    tmp = Path(tempfile.mkdtemp(prefix="detect_"))
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(path)],
        capture_output=True, text=True, check=True,
    )
    out = list(tmp.glob("*.xlsx"))
    if not out:
        raise RuntimeError(f"xls2xlsx failed for {path}")
    return out[0]


# ---- 核心探测 ----

def _scan_header_row(ws, max_scan_rows: int = 12) -> tuple[int, list[tuple[int, int, str]]] | None:
    """扫前 max_scan_rows 行，返回语义命中最多的行 + 该行命中列列表。"""
    best: tuple[int, list[tuple[int, int, str]]] | None = None
    best_count = 0
    max_col = min(ws.max_column, 40)
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        hits: list[tuple[int, int, str]] = []
        for c in range(1, max_col + 1):
            sem = _semantic_for(ws.cell(row=r, column=c).value)
            if sem:
                hits.append((r, c, sem))
        unique_sems = {h[2] for h in hits}
        if len(unique_sems) >= 4 and len(hits) > best_count:
            best = (r, hits)
            best_count = len(hits)
    return best


def _find_field_blocks(hits: list[tuple[int, int, str]]) -> list[list[tuple[int, str]]]:
    """根据命中列表，找连续的字段块（同行里相邻列语义序列，至少 3 个不同语义）。"""
    cols_sorted = sorted(hits, key=lambda h: h[1])
    blocks: list[list[tuple[int, str]]] = []
    current: list[tuple[int, str]] = []
    prev_col: int | None = None
    for _, col, sem in cols_sorted:
        if prev_col is None or col - prev_col <= 2:
            current.append((col, sem))
        else:
            if len({s for _, s in current}) >= 3:
                blocks.append(current)
            current = [(col, sem)]
        prev_col = col
    if current and len({s for _, s in current}) >= 3:
        blocks.append(current)
    return blocks


def _block_columns(block: list[tuple[int, str]]) -> dict[str, int]:
    result: dict[str, int] = {}
    for col, sem in block:
        result.setdefault(sem, col)
    return result


def _detect_data_start_row(ws, header_row: int) -> int:
    for r in range(header_row + 1, min(ws.max_row, header_row + 5) + 1):
        if any(_clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 20) + 1)):
            return r
    return header_row + 1


def _find_sap_side(ws, header_row: int, blocks: list[list[tuple[int, str]]]) -> int | None:
    """在 header_row-1 行找哪一块的起点处命中 SAP label。"""
    label_row = header_row - 1
    if label_row < 1:
        return None
    patterns = [re.compile(p, re.IGNORECASE) for p in SAP_LABEL_PATTERNS]
    left_label = _clean(ws.cell(row=label_row, column=1).value) or ""
    if any(p.search(left_label) for p in patterns):
        return 0
    for i, b in enumerate(blocks):
        start_col = b[0][0]
        lab = _clean(ws.cell(row=label_row, column=start_col).value) or ""
        if any(p.search(lab) for p in patterns):
            return i
    return None


def _find_aux_columns(ws, header_row: int, label_row: int | None = None) -> dict[str, int | None]:
    label_row = label_row or header_row - 1
    result: dict[str, int | None] = {k: None for k in AUX_VOCAB}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        t5 = _clean(ws.cell(row=max(1, label_row), column=c).value)
        t6 = _clean(ws.cell(row=header_row, column=c).value)
        joined = " ".join(str(x) for x in (t5, t6) if x)
        for key, vocab in AUX_VOCAB.items():
            if result[key] is None:
                if t6 and str(t6).strip() in vocab:
                    result[key] = c
                elif any(v in joined for v in vocab):
                    result[key] = c
    return result


def _is_mapping_sheet_name(name: str) -> bool:
    return any(k in name for k in MAPPING_SHEET_KEYWORDS)


def _regex_from_names(names: list[str]) -> str:
    """从实际工作表名推一个安全的匹配正则：MAPPING_SHEET_KEYWORDS 里命中所有 names 的最长关键词。"""
    if not names:
        return "項目マッピング"
    for kw in sorted(MAPPING_SHEET_KEYWORDS, key=len, reverse=True):
        if all(kw in n for n in names):
            return re.escape(kw)
    return "マッピング"


# ---- 探测主入口 ----

def detect_reference(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    candidates: list[tuple[str, int, list[list[tuple[int, str]]]]] = []
    for name in wb.sheetnames:
        if not _is_mapping_sheet_name(name):
            continue
        ws = wb[name]
        scan = _scan_header_row(ws)
        if not scan:
            continue
        hr, hits = scan
        blocks = _find_field_blocks(hits)
        if len(blocks) >= 2:
            candidates.append((name, hr, blocks))
    if not candidates:
        raise RuntimeError(f"no valid mapping sheet found in {xlsx_path.name}")
    candidates.sort(key=lambda x: -(len(x[2]) + sum(len({s for _, s in b}) for b in x[2])))
    sheet_name, header_row, blocks = candidates[0]
    ws = wb[sheet_name]

    left_block, right_block = blocks[0], blocks[1]
    sap_idx = _find_sap_side(ws, header_row, [left_block, right_block])
    sap_side = {0: "left", 1: "right", None: "right"}[sap_idx]

    return {
        "sheet_name_sample": sheet_name,
        "mapping_sheet_name_regex": _regex_from_names([s for s in wb.sheetnames if _is_mapping_sheet_name(s)]),
        "header_row": header_row,
        "data_start_row": _detect_data_start_row(ws, header_row),
        "label_row": max(1, header_row - 1),
        "left_block": _block_columns(left_block),
        "right_block": _block_columns(right_block),
        "sap_side": sap_side,
        "aux_columns": {k: v for k, v in _find_aux_columns(ws, header_row).items() if v},
    }


# ---- yaml 草稿输出 ----

def _dump_yaml(data: dict, header_comment: str = "") -> str:
    body = yaml.safe_dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False)
    return (header_comment + "\n" if header_comment else "") + body


def render_reference_yaml(detected: dict, project_name: str = "unknown") -> str:
    data = {
        "name": project_name,
        "description": "自动生成，请人工复核",
        "sources_dir": "sources",
        "out_dir": "knowledge",
        "mapping_sheet": {
            "name_regex": detected["mapping_sheet_name_regex"],
            "header_row": detected["header_row"],
            "data_start_row": detected["data_start_row"],
            "label_row": detected["label_row"],
        },
        "field_semantics": FIELD_SEMANTICS_VOCAB,
        "target_side": {"label_patterns": [r"部品\s*SAP", "ＳＡＰ"]},
        "aux_columns": AUX_VOCAB,
        "direction_rules": {
            "sheet_suffix": {"(受信)": "external_to_sap", "(返信)": "sap_to_external"},
            "file_marker":  {"※受信": "external_to_sap", "※送信": "sap_to_external"},
            "fallback_by_sap_side": True,
        },
        "if_meta": {"ifid_cell": [2, 1], "if_name_cell": [2, 3]},
    }
    header = (
        "# 自动探测生成 — 请人工复核\n"
        f"# 样本工作表：{detected['sheet_name_sample']}\n"
        f"# 检出字段块：left={detected['left_block']}\n"
        f"#           right={detected['right_block']}\n"
        f"# 推断 SAP 侧：{detected['sap_side']}\n"
        f"# [TODO] 请确认 if_meta.ifid_cell 与 if_name_cell 所在单元格\n"
    )
    return _dump_yaml(data, header)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="参考文档样本 Excel 路径")
    ap.add_argument("--project", default="unknown", help="项目名")
    ap.add_argument("--out", help="yaml 输出路径（默认打印到 stdout）")
    args = ap.parse_args()

    src = Path(args.xlsx).resolve()
    if not src.exists():
        print(f"not found: {src}", file=sys.stderr)
        return 1
    xlsx = _ensure_xlsx(src)

    detected = detect_reference(xlsx)
    yaml_text = render_reference_yaml(detected, args.project)

    if args.out:
        Path(args.out).write_text(yaml_text, encoding="utf-8")
        print(f"wrote → {args.out}")
    else:
        print(yaml_text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
