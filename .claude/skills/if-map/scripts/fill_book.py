"""为一份空白设计书从知识库查找 SAP 字段候选，直接回填到 Excel。

用法：
  python3 scripts/fill_book.py --project kaps2 <input.xls_or_xlsx>
  python3 scripts/fill_book.py --project kaps2 --schema <path.yaml> <input>

默认会在同目录寻找 <stem>.schema.yaml 作为输入格式描述。
产物：同目录下 <stem>_候选.xlsx（追加 3 列：推荐 SAP 字段 / 信心 / 说明）
  - "推荐 SAP 字段" 列带下拉菜单，可选 Top-3 或手动
  - 原 .xls 不修改；产物为 .xlsx（openpyxl 不支持写 .xls）

匹配信号分两类：
  [直接命中] L1 ext_tech 精确(1.0) | L2 ext_name 精确(0.7) | L3 ext_name 归一化(0.5)
  [推测]    S1 名称子串命中(0.4) | S2 上下文结构字典(0.4) | S3 业务词典(0.2)
  本书上下文 sap_struct 聚集 → +0.15（≥3 次聚集）
  推测类候选在 Excel 里显示前缀 "[推测]"，信心最高档位仅 ★（推测）。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
import json
import os
import subprocess
import sys
import tempfile
import unicodedata
import warnings
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 项目根：以调用方 CWD 为准（skill 从项目根运行）
ROOT = Path.cwd()
TOP_N = 3

# 追加列的起始位置（L 列 = 12）。若原表宽度已超过，动态推后。
DEFAULT_APPEND_START = 12

# 说明列颜色
FILL_NO_MATCH = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # 浅黄
FILL_STRONG   = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")   # 浅绿
FILL_WEAK     = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")   # 浅橙
FILL_HEADER   = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")   # 浅蓝

NO_MATCH_OPTION = "（无历史映射 — 可能为接口内部字段）"
SKIP_OPTION = "（填充/占位，跳过映射）"
SPECULATE_PREFIX = "[推测] "
AI_SPECULATE_PREFIX = "[AI推测] "

# AI 推测配置
AI_TIMEOUT_SEC = 120      # 单字段 AI 调用超时（claude -p 30-60s 是常态）
AI_MAX_HISTORY_HITS = 12  # 喂给 AI 的关键词历史命中条数上限
AI_MAX_STRUCT_FIELDS = 10 # 每个 top struct 下喂给 AI 的字段数上限

# 推测信号权重
SPEC_WEIGHT_SUBSTR = 0.15     # S1 名称子串（低权重 + freq 封顶，避免偶发 1 次主导）
SPEC_WEIGHT_DICT   = 0.60     # S3 业务词典（固定权重不乘 freq：语义优先于字面）
SPEC_SUBSTR_CAP    = 3        # S1 单个关键词 freq 封顶

# 名称过短不参与子串匹配
MIN_KEYWORD_LEN = 2

# 过于泛化的 token，不作为关键词
STOP_KEYWORDS = {
    "No", "NO", "no", "N0", "№", "ID", "ＩＤ", "コード", "CODE", "Code", "code",
    "番号", "のNo", "Flg", "FLG", "flag", "Flag", "FLAG", "status", "Status", "STATUS",
}

# L2 上下文加分（不独立产生候选）
CTX_STRUCT_BOOST = 0.35


# ------------- 基础工具 ------------- #

def _col_to_idx(spec: Any) -> int:
    if isinstance(spec, int):
        return spec
    s = str(spec).strip()
    if s.isdigit():
        return int(s)
    n = 0
    for ch in s.upper():
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"invalid column spec: {spec!r}")
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def normalize_name(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"[\s・\-_／/、。．\.№:：（）()\[\]【】]+", "", s)
    return s.lower()


def normalize_struct(s: str | None) -> str:
    """把 'LIPS\\nLIPS' 这类合并单元格残留归一化到单一 struct。"""
    if not s:
        return ""
    parts = [p.strip() for p in re.split(r"\n", str(s)) if p.strip()]
    return parts[0] if parts else ""


def normalize_multiline(s: str | None) -> str:
    """合并单元格多行 'NTGEW\\nBRGEW' → 'NTGEW+BRGEW'（人类可读地合并）。"""
    if not s:
        return ""
    parts = [p.strip() for p in re.split(r"\n", str(s)) if p.strip()]
    if len(parts) <= 1:
        return parts[0] if parts else ""
    # 去重保序
    seen = set()
    uniq = [p for p in parts if not (p in seen or seen.add(p))]
    return "+".join(uniq)


# ------------- .xls → .xlsx 自动转换 ------------- #

def ensure_xlsx(path: Path) -> Path:
    if path.suffix.lower() == ".xlsx":
        return path
    if path.suffix.lower() != ".xls":
        raise ValueError(f"unsupported extension: {path.suffix}")
    tmp = Path(tempfile.mkdtemp(prefix="fillbook_xls2xlsx_"))
    cmd = ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(path)]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr}")
    out = tmp / f"{path.stem}.xlsx"
    if not out.exists():
        candidates = list(tmp.glob("*.xlsx"))
        if not candidates:
            raise RuntimeError(f"no xlsx in {tmp}")
        out = candidates[0]
    return out


# ------------- schema & 字段抽取 ------------- #

def load_input_schema(input_path: Path, override: Path | None = None) -> dict:
    p = override if override else input_path.with_suffix(".schema.yaml")
    if not p.exists():
        p = input_path.parent / (input_path.stem + ".schema.yaml")
    if not p.exists():
        raise FileNotFoundError(
            f"input schema not found: expected {input_path.with_suffix('.schema.yaml')} or pass --schema"
        )
    with open(p, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _normalize_col_spec(spec: Any) -> int | list[int]:
    """支持 columns 的两种语法：
      - 单列：'C' / 3  → 列号 int
      - 多列（取第一个非空，用于 JSON 层级缩进字段）：['B','C','D']  → [2,3,4]
    """
    if isinstance(spec, list):
        return [_col_to_idx(s) for s in spec]
    return _col_to_idx(spec)


def _read_cell_multi(ws, r: int, col_spec: int | list[int]) -> Any:
    """按列号 / 列区间读取。多列时扫左到右取第一个非空值；同时返回命中深度（第几列，0-based）。"""
    if isinstance(col_spec, list):
        for depth, c in enumerate(col_spec):
            v = _clean(ws.cell(row=r, column=c).value)
            if v is not None:
                return v, depth
        return None, None
    v = _clean(ws.cell(row=r, column=col_spec).value)
    return v, 0 if v is not None else None


def read_blank_book(xlsx_path: Path, schema: dict) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = schema.get("sheet") or wb.sheetnames[0]
    ws = wb[sheet_name]

    data_start_row = schema["data_start_row"]
    cols_cfg = schema["columns"]
    col_idx = {sem: _normalize_col_spec(spec) for sem, spec in cols_cfg.items()}
    has_ext_no = "ext_no" in col_idx

    skip_names = set(schema.get("skip", {}).get("names") or [])
    skip_techs = set(schema.get("skip", {}).get("techs") or [])

    # 终止条件：连续 N 行 ext_name/ext_tech 都空就认为字段列表结束
    EMPTY_RUN_STOP = 4
    empty_run = 0
    auto_no = 0

    fields: list[dict] = []
    for r in range(data_start_row, ws.max_row + 1):
        rec: dict = {"row_idx": r}
        for sem, c in col_idx.items():
            val, depth = _read_cell_multi(ws, r, c)
            rec[sem] = val
            if sem == "ext_name" and depth is not None:
                rec["_depth"] = depth  # JSON 层级深度（0=最外层）
        if rec.get("ext_name") is None and rec.get("ext_tech") is None:
            empty_run += 1
            if empty_run >= EMPTY_RUN_STOP:
                break
            continue
        empty_run = 0

        # ext_no 逻辑：若 schema 声明且为非数字（页脚特征），停止；否则自动递增
        if has_ext_no:
            no_val = rec.get("ext_no")
            if no_val is None:
                continue
            if not (isinstance(no_val, (int, float)) or str(no_val).strip().isdigit()):
                break  # 末尾页脚
        else:
            auto_no += 1
            rec["ext_no"] = str(auto_no)

        name = rec.get("ext_name") or ""
        tech = rec.get("ext_tech") or ""
        if name in skip_names or tech in skip_techs:
            rec["skip"] = True
        fields.append(rec)
    return fields, schema.get("if_meta") or {}


# ------------- 候选检索 ------------- #

def query_by_ext_tech(kb: sqlite3.Connection, tech: str) -> list[sqlite3.Row]:
    return list(kb.execute("""
        SELECT sap_struct, sap_tech, sap_name,
               COUNT(*) AS freq,
               GROUP_CONCAT(DISTINCT ifid) AS ifs
        FROM field_mappings
        WHERE ext_tech = ? AND sap_tech IS NOT NULL
        GROUP BY sap_struct, sap_tech, sap_name
        ORDER BY freq DESC
    """, (tech,)))


def query_by_ext_name(kb: sqlite3.Connection, name: str) -> list[sqlite3.Row]:
    return list(kb.execute("""
        SELECT sap_struct, sap_tech, sap_name,
               COUNT(*) AS freq,
               GROUP_CONCAT(DISTINCT ifid) AS ifs
        FROM field_mappings
        WHERE ext_name = ? AND sap_tech IS NOT NULL
        GROUP BY sap_struct, sap_tech, sap_name
        ORDER BY freq DESC
    """, (name,)))


def count_ext_tech_in_kb(kb: sqlite3.Connection, tech: str) -> int:
    """历史里 ext_tech 出现总次数（含 SAP 侧为空的行）。"""
    row = kb.execute("SELECT COUNT(*) FROM field_mappings WHERE ext_tech = ?", (tech,)).fetchone()
    return row[0] if row else 0


def build_norm_index(kb: sqlite3.Connection) -> dict[str, set[str]]:
    idx: dict[str, set[str]] = defaultdict(set)
    for (name,) in kb.execute("SELECT DISTINCT ext_name FROM field_mappings WHERE ext_name IS NOT NULL"):
        idx[normalize_name(name)].add(name)
    return idx


# ------------- 推测层辅助：业务词典 / 结构字典 / 关键词提取 ------------- #

def load_business_dict(project_dir: Path) -> dict:
    """合并加载业务语义词典：默认（scripts/business_dict.default.yaml）+ 项目（projects/<name>/business_dict.yaml）。

    合并策略：项目词典在前，默认在后（项目 regex 优先命中）。
    """
    merged_patterns: list[dict] = []

    proj = project_dir / "business_dict.yaml"
    if proj.exists():
        with open(proj, encoding="utf-8") as f:
            d = yaml.safe_load(f) or {}
            merged_patterns.extend(d.get("patterns") or [])

    default = Path(__file__).resolve().parent / "business_dict.default.yaml"
    if default.exists():
        with open(default, encoding="utf-8") as f:
            d = yaml.safe_load(f) or {}
            default_patterns = d.get("patterns") or []
            # 去重：项目已有同 regex 的不再加
            existing_regexes = {p.get("regex") for p in merged_patterns}
            for p in default_patterns:
                if p.get("regex") not in existing_regexes:
                    merged_patterns.append(p)

    return {"patterns": merged_patterns}


def build_struct_field_dict(kb: sqlite3.Connection) -> dict[str, list[dict]]:
    """每个 sap_struct 下历史出现过的字段清单（按频次排序）。"""
    result: dict[str, list[dict]] = defaultdict(list)
    rows = kb.execute("""
        SELECT sap_struct, sap_tech, sap_name,
               COUNT(*) AS freq,
               GROUP_CONCAT(DISTINCT ifid) AS ifs
        FROM field_mappings
        WHERE sap_tech IS NOT NULL AND sap_struct IS NOT NULL
        GROUP BY sap_struct, sap_tech, sap_name
    """).fetchall()
    for r in rows:
        struct = normalize_struct(r["sap_struct"])
        tech = normalize_multiline(r["sap_tech"])
        name = normalize_multiline(r["sap_name"])
        result[struct].append({
            "sap_struct": struct, "sap_tech": tech, "sap_name": name,
            "freq": r["freq"], "ifs": set((r["ifs"] or "").split(",")) - {""},
        })
    for s in result:
        result[s].sort(key=lambda x: -x["freq"])
    return dict(result)


# 日文/中文 token 识别
_TOKEN_RE = re.compile(r"[一-鿿぀-ゟ゠-ヿ]+|[A-Za-z]{2,}")
# 常见后缀（会单独提取，通常触发语义词典）
_COMMON_SUFFIXES = ("日付", "年月日", "時刻", "時間", "コード", "数量", "重量", "金額", "区分", "形態", "種類")
_STRIP_SUFFIXES = ("コード", "№", "No", "no", "NO", "数", "日")


def extract_keywords(name: str | None) -> list[str]:
    """从字段名提取用于"名称子串匹配"的关键词。

    策略：
      1. 取整词 tokens（含 ≥2 字符的日文/中文片段 或 ≥2 字母）
      2. 再加"去掉常见后缀后的词根"作为另一个关键词
    避免 1-gram（太泛化）和太短 token。
    """
    if not name:
        return []
    s = unicodedata.normalize("NFKC", name)
    # 去圆括号补充说明
    s = re.sub(r"[（(][^）)]*[）)]", "", s)
    tokens = _TOKEN_RE.findall(s)
    kws: list[str] = []
    seen: set[str] = set()

    def add(tok: str):
        if not tok or len(tok) < MIN_KEYWORD_LEN:
            return
        if tok in STOP_KEYWORDS:
            return
        if tok in seen:
            return
        seen.add(tok)
        kws.append(tok)

    for t in tokens:
        add(t)
        # 去后缀的词根
        for suf in _STRIP_SUFFIXES:
            if len(t) > len(suf) + 1 and t.endswith(suf):
                add(t[: -len(suf)])
                break
    return kws


def check_skip_patterns(name: str | None, business_dict: dict) -> str | None:
    """检查 ext_name 是否命中业务词典里的 skip_reason。命中则返回 reason 字符串。"""
    if not name:
        return None
    for pat in business_dict.get("patterns") or []:
        if "skip_reason" in pat and re.search(pat["regex"], name):
            return pat["skip_reason"]
    return None


# ------------- 复合字段拆分 ------------- #

_COMPOSITE_RE = re.compile(
    r"(?P<a>\S+?)\s*(?P<la>\d+)\s*(?:桁|文字|byte|Byte|位|chars?)\s*\+\s*"
    r"(?P<b>\S+?)\s*(?P<lb>\d+)\s*(?:桁|文字|byte|Byte|位|chars?)"
)


def split_composite(field: dict) -> list[dict] | None:
    """若备注里有 'X6桁+Y3桁' 形式的复合指示，拆成两个虚拟子字段。"""
    remark = field.get("remark") or ""
    m = _COMPOSITE_RE.search(unicodedata.normalize("NFKC", str(remark)))
    if not m:
        return None
    return [
        {**field, "ext_name": m.group("a").strip(), "ext_len": m.group("la"),
         "_composite_parent": True, "_composite_label": m.group("a").strip()},
        {**field, "ext_name": m.group("b").strip(), "ext_len": m.group("lb"),
         "_composite_parent": True, "_composite_label": m.group("b").strip()},
    ]


def _new_cand(struct_n, tech_n, name_n):
    return {
        "sap_struct": struct_n, "sap_tech": tech_n, "sap_name": name_n,
        "signals": {}, "weighted_freq": 0.0, "raw_freq": 0,
        "ifs": set(), "origin": "history",
    }


def _add_cand(agg, sap_struct, sap_tech, sap_name, freq, ifs, hit_type, weight, origin="history"):
    struct_n = normalize_struct(sap_struct)
    tech_n = normalize_multiline(sap_tech)
    name_n = normalize_multiline(sap_name)
    if not tech_n:
        return
    key = (struct_n, tech_n)
    if key not in agg:
        agg[key] = _new_cand(struct_n, tech_n, name_n)
        agg[key]["origin"] = origin
    elif origin == "history":
        agg[key]["origin"] = "history"  # history 胜过 speculate
    c = agg[key]
    c["signals"][hit_type] = c["signals"].get(hit_type, 0) + freq
    c["weighted_freq"] += freq * weight
    c["raw_freq"] += freq
    if ifs:
        if isinstance(ifs, str):
            c["ifs"].update(s for s in ifs.split(",") if s)
        else:
            c["ifs"].update(ifs)


def pass1_candidates(field: dict, kb: sqlite3.Connection, norm_index: dict) -> list[dict]:
    agg: dict[tuple, dict] = {}

    if field.get("ext_tech"):
        for r in query_by_ext_tech(kb, field["ext_tech"]):
            _add_cand(agg, r["sap_struct"], r["sap_tech"], r["sap_name"], r["freq"], r["ifs"],
                      "tech_exact", 1.0, origin="history")

    if field.get("ext_name"):
        for r in query_by_ext_name(kb, field["ext_name"]):
            _add_cand(agg, r["sap_struct"], r["sap_tech"], r["sap_name"], r["freq"], r["ifs"],
                      "name_exact", 0.7, origin="history")

        nq = normalize_name(field["ext_name"])
        hit_names = norm_index.get(nq, set()) - {field["ext_name"]}
        for hn in hit_names:
            for r in query_by_ext_name(kb, hn):
                _add_cand(agg, r["sap_struct"], r["sap_tech"], r["sap_name"], r["freq"], r["ifs"],
                          "name_norm", 0.5, origin="history")

    return sorted(agg.values(), key=lambda x: x["weighted_freq"], reverse=True)


def pass1_speculate(
    field: dict, kb: sqlite3.Connection,
    business_dict: dict, struct_dict: dict[str, list[dict]], ctx_structs: list[str],
) -> list[dict]:
    """无直接命中时的推测层：
      S1 名称子串（独立产出候选）
      S3 业务词典（独立产出候选）
      S2 上下文结构只对已产出候选做加分（避免把 VBAP.MATNR 这类高频主键刷到所有字段）
    """
    agg: dict[tuple, dict] = {}
    name = field.get("ext_name") or ""

    # S1 名称子串（freq 封顶 + 同一 SAP 候选按最高关键词命中只计一次，避免多关键词叠加）
    substr_hits: dict[tuple, dict] = {}
    for kw in extract_keywords(name):
        rows = kb.execute("""
            SELECT sap_struct, sap_tech, sap_name,
                   COUNT(*) AS freq, GROUP_CONCAT(DISTINCT ifid) AS ifs
            FROM field_mappings
            WHERE ext_name LIKE ? AND sap_tech IS NOT NULL
            GROUP BY sap_struct, sap_tech, sap_name
        """, (f"%{kw}%",)).fetchall()
        for r in rows:
            struct_n = normalize_struct(r["sap_struct"])
            tech_n = normalize_multiline(r["sap_tech"])
            key = (struct_n, tech_n)
            capped = min(r["freq"], SPEC_SUBSTR_CAP)
            existing = substr_hits.get(key)
            if existing is None or existing["capped"] < capped:
                substr_hits[key] = {
                    "capped": capped, "kw": kw, "sap_struct": r["sap_struct"],
                    "sap_tech": r["sap_tech"], "sap_name": r["sap_name"], "ifs": r["ifs"],
                }
    for hit in substr_hits.values():
        _add_cand(agg, hit["sap_struct"], hit["sap_tech"], hit["sap_name"],
                  hit["capped"], hit["ifs"], f"substr({hit['kw']})",
                  SPEC_WEIGHT_SUBSTR, origin="speculate")

    # S3 业务词典（固定权重不乘 freq，语义优先于字面）
    for pat in business_dict.get("patterns") or []:
        if "skip_reason" in pat:
            continue
        if re.search(pat["regex"], name):
            # 按 suggest 的排序顺序降权：第 1 个 × 1.0，第 2 个 × 0.7，第 3 个 × 0.5
            for idx, sug in enumerate(pat.get("suggest") or []):
                rank_factor = [1.0, 0.7, 0.5][idx] if idx < 3 else 0.3
                _add_cand(agg, sug["struct"], sug["tech"], sug.get("name"),
                          1, None, f"dict({pat['regex'][:20]})",
                          SPEC_WEIGHT_DICT * rank_factor, origin="speculate")

    # S2 上下文结构加分（仅对已聚合候选）
    ctx_top = set(ctx_structs[:3])
    for c in agg.values():
        if c["sap_struct"] in ctx_top:
            c["weighted_freq"] += CTX_STRUCT_BOOST
            c["signals"]["ctx_boost"] = c["signals"].get("ctx_boost", 0) + 1

    return sorted(agg.values(), key=lambda x: x["weighted_freq"], reverse=True)


def _build_ai_prompt(
    field: dict,
    kb: sqlite3.Connection,
    business_dict: dict,
    struct_dict: dict[str, list[dict]],
    ctx_counter: Counter,
    ctx_structs: list[str],
    book_field_summary: list[dict],
    if_meta: dict,
) -> str:
    """构建喂给 AI 的上下文 prompt：字段信息 + 业务场景 + 本书上下文 + 检索辅助 + 任务规约。"""
    # ---- 1. 字段信息 ----
    name = field.get("ext_name") or ""
    tech = field.get("ext_tech") or ""
    field_lines = [
        f"项目名: {name}",
        f"项目代码: {tech}",
        f"类型: {field.get('ext_type') or '?'}  长度: {field.get('ext_len') or '?'}  字节: {field.get('ext_byte') or '?'}",
        f"备注: {field.get('remark') or '(无)'}",
    ]

    # ---- 2. 本接口业务场景 ----
    ifid = if_meta.get("ifid_guess") or "?"
    if_name = if_meta.get("if_name") or "?"
    direction = if_meta.get("direction") or "unknown"
    direction_zh = {
        "external_to_sap": "外部系统 → SAP（入向）",
        "sap_to_external": "SAP → 外部系统（出向）",
    }.get(direction, direction)
    counterpart = if_meta.get("counterpart_hint") or "?"
    biz_lines = [
        f"IFID: {ifid}",
        f"业务名: {if_name}",
        f"方向: {direction_zh}",
        f"对向系统: {counterpart}",
    ]

    # ---- 3. 本书已确定的字段分布 ----
    resolved = [f for f in book_field_summary if f.get("sap_struct")]
    skipped = [f for f in book_field_summary if f.get("is_skipped")]
    by_struct: dict[str, list[str]] = {}
    for r in resolved:
        s = r["sap_struct"]
        by_struct.setdefault(s, []).append(
            f"{r['ext_name']}({r['ext_tech']}) → {s}.{r['sap_tech']}（{r.get('sap_name') or ''}）"
        )
    book_lines = []
    for s, entries in sorted(by_struct.items(), key=lambda x: -len(x[1])):
        book_lines.append(f"- {s} ({len(entries)} 字段):")
        for e in entries[:6]:
            book_lines.append(f"    · {e}")
    if skipped:
        names = ", ".join((f"{f.get('ext_name')}({f.get('ext_tech')})" for f in skipped[:8]))
        book_lines.append(f"- 已跳过接口内部字段: {names}")
    if not book_lines:
        book_lines.append("(本书其他字段尚未确定 SAP 映射)")

    # ---- 4. 检索辅助 ----
    # 4a. 关键词历史命中
    kw_hits: list[str] = []
    for kw in extract_keywords(name):
        rows = kb.execute("""
            SELECT ext_name, sap_struct, sap_tech, sap_name, COUNT(*) AS freq
            FROM field_mappings
            WHERE ext_name LIKE ? AND sap_tech IS NOT NULL
            GROUP BY ext_name, sap_struct, sap_tech, sap_name
            ORDER BY freq DESC LIMIT 4
        """, (f"%{kw}%",)).fetchall()
        for r in rows:
            struct_n = normalize_struct(r["sap_struct"])
            tech_n = normalize_multiline(r["sap_tech"])
            kw_hits.append(
                f"[关键词 '{kw}'] {r['ext_name']} → {struct_n}.{tech_n}（{r['sap_name'] or ''}）×{r['freq']}"
            )
            if len(kw_hits) >= AI_MAX_HISTORY_HITS:
                break
        if len(kw_hits) >= AI_MAX_HISTORY_HITS:
            break

    # 4b. 业务词典候选
    dict_hits: list[str] = []
    dict_skip_hit: str | None = None
    for pat in business_dict.get("patterns") or []:
        if not re.search(pat["regex"], name):
            continue
        if "skip_reason" in pat:
            dict_skip_hit = pat["skip_reason"]
            continue
        for sug in pat.get("suggest") or []:
            dict_hits.append(
                f"[规则 '{pat['regex'][:20]}'] → {sug['struct']}.{sug['tech']}（{sug.get('name') or ''}）"
            )
        if pat.get("hint"):
            dict_hits.append(f"  ↳ 业务提示: {pat['hint']}")

    # 4c. 本书 Top 结构下的常见字段
    struct_field_lines: list[str] = []
    for s in ctx_structs[:3]:
        entries = struct_dict.get(s, [])[:AI_MAX_STRUCT_FIELDS]
        if not entries:
            continue
        struct_field_lines.append(f"- {s} 常见字段:")
        for e in entries:
            struct_field_lines.append(f"    · {s}.{e['sap_tech']}（{e.get('sap_name') or ''}）")

    # ---- 组装 prompt ----
    parts = [
        "你是 SAP 接口设计顾问。帮我为一个外部系统字段推荐对应的 SAP 表和字段。",
        "",
        "## 待推荐字段",
        *field_lines,
        "",
        "## 本接口业务场景",
        *biz_lines,
        "",
        "## 本接口已确定的字段分布（用于推断业务上下文）",
        *book_lines,
    ]
    if kw_hits:
        parts += ["", "## 知识库中按关键词检索到的历史映射（参考）", *kw_hits]
    if dict_hits:
        parts += ["", "## 业务词典规则命中的候选", *dict_hits]
    if dict_skip_hit:
        parts += ["", f"## 业务词典的 skip 提示", f"- {dict_skip_hit}"]
    if struct_field_lines:
        parts += ["", "## 本书上下文 Top 结构的常见字段（知识库统计）", *struct_field_lines]

    parts += [
        "",
        "## 类型与长度约束（重要 — 候选字段必须类型兼容）",
        "- C(1-5) 短字符：通常为**代码/区分/フラグ/タイプ/状态**；映到 SAP 的 CHAR 短字段"
        "（如 LVORM、SPART、AUART、WERKS、LFART 等）。**严禁**映到 QUAN 数量字段"
        "（LFIMG/KWMENG/MENGE 精度 13+3）——哪怕备注写『PIC \"999\" 符号なし』，"
        "SAP 数量字段为精确数值设计；短 C 字段表示的『数』多为 API 聚合的件数统计，应判 skip。",
        "- C(6-10)：多为编码/ID（注番、品番、伝票番号、得意先コード）。",
        "- C(>10)：多为名称/说明/URL/描述文本。",
        "- C(6) 且备注含 YYMMDD、或 C(8) 含 YYYYMMDD：日期，SAP 用 DATS(8)。",
        "- C(6) 且备注含 HHMMSS：时刻，SAP 用 TIMS(6)。",
        "- 真正的『数量/金額』：SAP 是 QUAN/CURR（精度大），要求接口侧用足够精度的类型（通常 P 型或长 N 型），"
        "短 C 字段不兼容；此时宁可 skip=true（判为 API 聚合计数值）也不要强塞 LFIMG/KWMENG。",
        "- 若所有候选类型/长度都不兼容，返回 skip=true 并说明原因。",
        "",
        "## 任务",
        "基于以上全部信息（包括类型/长度约束），给这个字段推荐 Top-3 SAP 候选，或判定为接口内部字段不需映射。",
        "每个候选的 reason 里**必须**解释为什么类型/长度兼容，不要只讲业务语义。",
        "只能使用真实存在的 SAP 标准表字段；不要编造。",
        "严格返回 JSON（不要其他文字），格式：",
        '''{
  "skip": false,
  "reason": "(若 skip=true，解释为何不映射；否则留空)",
  "candidates": [
    {
      "sap_struct": "LIPS",
      "sap_tech":   "LFIMG",
      "sap_name":   "出荷数量",
      "confidence": 0.85,
      "reason":     "本接口在讨论出荷（LIKP/LIPS 家族），'数' 后缀 + 出荷上下文 → 出荷数量"
    }
  ]
}''',
    ]
    return "\n".join(parts)


def _call_claude_p(prompt: str, timeout: int = AI_TIMEOUT_SEC) -> dict | None:
    """调 claude -p（子进程）。成功返回解析后的 JSON dict；失败返回 None。"""
    try:
        env = os.environ.copy()
        # 防止嵌套 claude -p 时环境变量干扰
        env.pop("CLAUDECODE", None)
        res = subprocess.run(
            ["claude", "-p"],
            input=prompt,
            capture_output=True, text=True, timeout=timeout,
            env=env,
        )
    except (subprocess.TimeoutExpired, OSError):
        return None
    if res.returncode != 0:
        return None
    # 从 stdout 里提取第一个完整 JSON（可能被 markdown 代码块包裹）
    out = res.stdout
    # 去掉 ```json … ``` 包裹
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", out, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    # 再尝试裸 JSON
    m = re.search(r"\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}", out, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            return None
    return None


def ai_speculate(
    field: dict,
    kb: sqlite3.Connection,
    business_dict: dict,
    struct_dict: dict[str, list[dict]],
    ctx_counter: Counter,
    ctx_structs: list[str],
    book_field_summary: list[dict],
    if_meta: dict,
) -> tuple[list[dict], str | None]:
    """让 AI 做推测。返回 (候选列表, skip_reason)。
    - 列表非空：AI 给出候选（origin="ai_speculate"）
    - skip_reason 非空：AI 判定为接口内部字段
    - 两者都空：AI 调用失败（调用方应降级到规则推测）
    """
    prompt = _build_ai_prompt(
        field, kb, business_dict, struct_dict, ctx_counter, ctx_structs,
        book_field_summary, if_meta,
    )
    result = _call_claude_p(prompt)
    if result is None:
        return [], None  # AI 失败，调用方降级
    if result.get("skip"):
        return [], result.get("reason") or "AI 判定为接口内部字段"
    cands: list[dict] = []
    for c in result.get("candidates") or []:
        struct = str(c.get("sap_struct") or "").strip()
        tech = str(c.get("sap_tech") or "").strip()
        if not struct or not tech:
            continue
        try:
            conf = float(c.get("confidence", 0.5))
        except (TypeError, ValueError):
            conf = 0.5
        cands.append({
            "sap_struct": struct,
            "sap_tech": tech,
            "sap_name": c.get("sap_name") or "",
            "signals": {"ai_speculate": 1},
            "weighted_freq": max(0.0, min(1.0, conf)),
            "raw_freq": 1,
            "ifs": set(),
            "origin": "ai_speculate",
            "ai_reason": c.get("reason") or "",
            "ai_confidence": max(0.0, min(1.0, conf)),
        })
    return cands[:TOP_N], None


def _build_book_field_summary(
    fields: list[dict], direct_pairs: list[tuple[dict, list[dict]]],
    business_dict: dict,
) -> list[dict]:
    """为 AI 准备本书字段分布摘要。"""
    summary: list[dict] = []
    direct_map = {id(f): cands for f, cands in direct_pairs}
    for f in fields:
        entry = {
            "ext_name": f.get("ext_name"),
            "ext_tech": f.get("ext_tech"),
            "sap_struct": None,
            "sap_tech": None,
            "sap_name": None,
            "is_skipped": False,
        }
        if f.get("skip"):
            entry["is_skipped"] = True
            summary.append(entry)
            continue
        cands = direct_map.get(id(f)) or []
        if cands:
            top = cands[0]
            entry["sap_struct"] = top["sap_struct"]
            entry["sap_tech"] = top["sap_tech"]
            entry["sap_name"] = top.get("sap_name")
        elif check_skip_patterns(f.get("ext_name"), business_dict):
            entry["is_skipped"] = True
        summary.append(entry)
    return summary


def resolve_candidates(
    field: dict, kb: sqlite3.Connection, norm_index: dict,
    business_dict: dict, struct_dict: dict, ctx_structs: list[str],
    # 以下为 AI 推测所需的额外上下文（向后兼容可选）
    ctx_counter: Counter | None = None,
    book_field_summary: list[dict] | None = None,
    if_meta: dict | None = None,
    enable_ai: bool = True,
) -> tuple[list[dict], str | None]:
    """统一入口：L0 直接命中 → 若无 → AI 推测（若启用）→ 失败降级到规则推测。"""
    if field.get("skip"):
        return [], "填充/占位字段，无需映射。"

    skip_reason = check_skip_patterns(field.get("ext_name"), business_dict)
    direct = pass1_candidates(field, kb, norm_index)
    if direct:
        return direct, None
    if skip_reason:
        return [], skip_reason

    # AI 推测（首选）
    if enable_ai and book_field_summary is not None and if_meta is not None:
        ai_cands, ai_skip = ai_speculate(
            field, kb, business_dict, struct_dict,
            ctx_counter or Counter(), ctx_structs,
            book_field_summary, if_meta,
        )
        if ai_skip:
            return [], f"[AI 判断] {ai_skip}"
        if ai_cands:
            return ai_cands, None
        # AI 失败 → 降级到规则

    spec = pass1_speculate(field, kb, business_dict, struct_dict, ctx_structs)
    spec = _diversify_speculate(spec, limit=TOP_N)
    return spec, None


def _diversify_speculate(cands: list[dict], limit: int = 3) -> list[dict]:
    """推测候选去重：每个 sap_struct 最多取一个代表，避免同结构刷屏。"""
    seen_structs: set[str] = set()
    out: list[dict] = []
    for c in cands:
        if c["sap_struct"] in seen_structs:
            continue
        seen_structs.add(c["sap_struct"])
        out.append(c)
        if len(out) >= limit:
            break
    return out


def context_structs(pairs: list[tuple[dict, list[dict]]]) -> Counter:
    c: Counter = Counter()
    for field, cands in pairs:
        if field.get("skip") or not cands:
            continue
        c[cands[0]["sap_struct"]] += 1
    return c


def pass2_score(cand: dict, total_weighted: float, ctx: Counter) -> float:
    base = cand["weighted_freq"] / total_weighted if total_weighted else 0.0
    bonus = 0.15 if ctx.get(cand["sap_struct"], 0) >= 3 else 0.0
    return min(1.0, base + bonus)


# ------------- 业务化说明文案 ------------- #

def confidence_label(score: float, raw_freq_top: int, origin: str = "history") -> str:
    # AI 推测封顶 ★★（AI推测），高于规则推测一档（AI 综合语义 + 上下文判断）
    if origin == "ai_speculate":
        if score >= 0.8:
            return "★★（AI推测）"
        if score >= 0.5:
            return "★（AI推测）"
        return "AI推测 需复核"
    # 规则推测封顶 ★（推测）
    if origin == "speculate":
        return "★（推测）"
    if raw_freq_top <= 1 and score >= 0.85:
        return "★★（历史仅1次，需复核）"
    if score >= 0.85:
        return "★★★"
    if score >= 0.6:
        return "★★"
    if score >= 0.3:
        return "★"
    return "需确认"


def cand_label(c: dict) -> str:
    """候选在下拉/推荐列的显示字符串。"""
    name = c.get("sap_name") or ""
    core = f"{c['sap_struct']}.{c['sap_tech']}（{name}）" if name else f"{c['sap_struct']}.{c['sap_tech']}"
    if c.get("origin") == "ai_speculate":
        return AI_SPECULATE_PREFIX + core
    if c.get("origin") == "speculate":
        return SPECULATE_PREFIX + core
    return core


def explain_no_match(field: dict, kb: sqlite3.Connection) -> str:
    """无候选时的说明，尽量根据历史数据给出具体结论。"""
    tech = field.get("ext_tech")
    name = field.get("ext_name") or "?"
    if tech:
        hist = count_ext_tech_in_kb(kb, tech)
        if hist > 0:
            return (
                f"在历史 {hist} 份映射中出现过"
                f"「{name}（{tech}）」，但都没有对应到 SAP 表字段。"
                f"推测为接口控制字段（文件ID/序号/状态码 等），不需映射。"
            )
    return (
        f"历史中没有「{name}」的 SAP 映射记录。可能是本接口新增字段或接口内部控制信息；"
        f"若确属业务字段，请人工指定对应的 SAP 字段。"
    )


def explain_matched(
    cands: list[dict], total_w: float, ctx: Counter, top_score: float
) -> str:
    top = cands[0]
    # AI 推测走独立文案
    if top.get("origin") == "ai_speculate":
        return _explain_ai_speculated(cands)
    # 规则推测走独立文案
    if top.get("origin") == "speculate":
        return _explain_speculated(cands)

    top_name = top.get("sap_name") or top["sap_tech"]
    top_loc = f"{top['sap_struct']}.{top['sap_tech']}"
    ifs_ref = sorted(top["ifs"])[:3]
    ifs_str = "、".join(ifs_ref)

    ctx_hit = ctx.get(top["sap_struct"], 0) >= 3

    if len(cands) == 1:
        return (
            f"历史上一致映射到「{top_name}」（{top_loc}），"
            f"共 {top['raw_freq']} 次，参考：{ifs_str}。建议直接采用。"
        )

    top2 = cands[1]
    top_share = top["weighted_freq"] / total_w if total_w else 0.0
    t2_name = top2.get("sap_name") or top2["sap_tech"]
    t2_loc = f"{top2['sap_struct']}.{top2['sap_tech']}"

    if top_share >= 0.7:
        return (
            f"多数历史映射到「{top_name}」（{top_loc}），"
            f"少数映射到「{t2_name}」（{t2_loc}）。"
            f"建议采用主选项，除非业务场景特殊。"
            + (f" 本书多个字段已指向 {top['sap_struct']}，该推荐更可信。" if ctx_hit else "")
        )
    elif abs(top["weighted_freq"] - top2["weighted_freq"]) / max(total_w, 1) < 0.15:
        return (
            f"历史上存在多种映射方式，「{top_name}」（{top_loc}）与"
            f"「{t2_name}」（{t2_loc}）频次相近。"
            f"两个 SAP 字段语义不同，请根据实际业务判断应使用哪一个。"
        )
    else:
        return (
            f"主要映射是「{top_name}」（{top_loc}），也有若干映射到"
            f"「{t2_name}」（{t2_loc}）。请结合业务判断。"
        )


def _explain_speculated(cands: list[dict]) -> str:
    """推测类候选的说明文案 —— 注明信号来源、限定"推测"。"""
    top = cands[0]
    top_name = top.get("sap_name") or top["sap_tech"]
    top_loc = f"{top['sap_struct']}.{top['sap_tech']}"
    sigs = list(top.get("signals", {}).keys())
    # 信号来源的自然语言
    src_parts: list[str] = []
    if any(s.startswith("substr(") for s in sigs):
        kws = [s[7:-1] for s in sigs if s.startswith("substr(")]
        src_parts.append(f"字段名含关键词 {'/'.join(kws[:3])}")
    if any(s.startswith("ctx_struct(") for s in sigs):
        structs = [s[11:-1] for s in sigs if s.startswith("ctx_struct(")]
        src_parts.append(f"本书上下文常用 {'/'.join(set(structs))} 表")
    if any(s.startswith("dict(") for s in sigs):
        src_parts.append("业务词典规则命中")
    src_str = "；".join(src_parts) if src_parts else "综合推断"

    msg = (
        f"历史中无直接映射，基于「{src_str}」推测对应到「{top_name}」（{top_loc}）。"
    )
    if len(cands) > 1:
        t2 = cands[1]
        msg += f" 备选：{t2['sap_struct']}.{t2['sap_tech']}（{t2.get('sap_name') or ''}）。"
    msg += " 请人工复核。"
    return msg


def _explain_ai_speculated(cands: list[dict]) -> str:
    """AI 推测候选的说明文案：展示 AI 给的 reason + 候选列表，强调需人工复核。"""
    top = cands[0]
    top_name = top.get("sap_name") or top["sap_tech"]
    top_loc = f"{top['sap_struct']}.{top['sap_tech']}"
    top_reason = top.get("ai_reason") or ""
    top_conf = top.get("ai_confidence", top.get("weighted_freq", 0.0))

    msg = f"AI 推测（无历史直接命中，信心 {top_conf:.2f}）：推荐「{top_name}」（{top_loc}）"
    if top_reason:
        msg += f"——{top_reason}"
    if len(cands) > 1:
        alt_bits = []
        for c in cands[1:TOP_N]:
            name = c.get('sap_name') or ''
            alt_bits.append(f"{c['sap_struct']}.{c['sap_tech']}（{name}）")
        msg += f"。备选：{' / '.join(alt_bits)}"
    msg += "。AI 推断仅作参考，请务必人工复核。"
    return msg


# ------------- Excel 回填 ------------- #

def render_excel(
    xlsx_path: Path, fields: list[dict],
    results: list[dict],  # [{"field":f, "entries":[...], "composite":bool}]
    ctx: Counter, schema: dict, out_path: Path, kb: sqlite3.Connection,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = schema.get("sheet") or wb.sheetnames[0]
    ws = wb[sheet_name]

    start_col = max(DEFAULT_APPEND_START, ws.max_column + 1)
    header_row = schema["header_row"]

    headers = ["推奨 SAP 字段", "信心", "備考（业务判断）"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=h)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate([40, 18, 60]):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w

    for item in results:
        f = item["field"]
        entries = item["entries"]  # 每项: (sub_field_label, cands, total_w, skip_reason)
        r = f["row_idx"]

        if f.get("skip"):
            ws.cell(row=r, column=start_col, value=SKIP_OPTION)
            ws.cell(row=r, column=start_col + 1, value="—")
            ws.cell(row=r, column=start_col + 2, value="填充/占位字段，无需映射。")
            _fill_row(ws, r, start_col, FILL_NO_MATCH)
            continue

        # 复合字段：把子字段的结果串起来
        if item.get("composite") and entries:
            pieces: list[str] = []
            expl_pieces: list[str] = []
            any_spec = False
            any_low = False
            for sub_label, cands, total_w, skip in entries:
                if skip:
                    pieces.append(f"{sub_label}→[跳过]")
                    expl_pieces.append(f"{sub_label}：{skip}")
                    continue
                if not cands:
                    pieces.append(f"{sub_label}→[无推荐]")
                    expl_pieces.append(f"{sub_label}：无可用推荐，请人工指定")
                    any_low = True
                    continue
                top = cands[0]
                pieces.append(f"{sub_label}→{cand_label(top)}")
                score = pass2_score(top, total_w, ctx)
                expl_pieces.append(
                    f"{sub_label}（{score:.2f}）：" +
                    (_explain_speculated([top]) if top.get("origin") == "speculate"
                     else explain_matched(cands, total_w, ctx, score))
                )
                if top.get("origin") == "speculate":
                    any_spec = True
                if score < 0.6:
                    any_low = True
            ws.cell(row=r, column=start_col, value=" ＋ ".join(pieces))
            ws.cell(row=r, column=start_col + 1, value="复合字段").alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=start_col + 2, value="本字段由多个子字段拼接，下列对每个子字段分别给出推荐：\n" + "\n".join(expl_pieces)).alignment = Alignment(wrap_text=True, vertical="top")
            fill = FILL_WEAK if (any_spec or any_low) else FILL_STRONG
            _fill_row(ws, r, start_col, fill)
            continue

        # 普通字段（单一）
        _, cands, total_w, skip_reason = entries[0]

        if skip_reason and not cands:
            ws.cell(row=r, column=start_col, value=NO_MATCH_OPTION)
            ws.cell(row=r, column=start_col + 1, value="—")
            ws.cell(row=r, column=start_col + 2, value=f"业务词典判定：{skip_reason}")
            _fill_row(ws, r, start_col, FILL_NO_MATCH)
            _attach_dv(ws, r, start_col, [NO_MATCH_OPTION, "（手动指定）"])
            continue

        if not cands:
            ws.cell(row=r, column=start_col, value=NO_MATCH_OPTION)
            ws.cell(row=r, column=start_col + 1, value="—")
            ws.cell(row=r, column=start_col + 2, value=explain_no_match(f, kb))
            _fill_row(ws, r, start_col, FILL_NO_MATCH)
            _attach_dv(ws, r, start_col, [NO_MATCH_OPTION, "（手动指定）"])
            continue

        top3 = cands[:TOP_N]
        labels = [cand_label(c) for c in top3]
        top_score = pass2_score(top3[0], total_w, ctx)
        stars = confidence_label(top_score, top3[0]["raw_freq"], top3[0].get("origin", "history"))
        explanation = explain_matched(top3, total_w, ctx, top_score)

        ws.cell(row=r, column=start_col, value=labels[0])
        ws.cell(row=r, column=start_col + 1, value=stars).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=start_col + 2, value=explanation).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

        if top3[0].get("origin") == "speculate":
            fill = FILL_WEAK
        elif top_score >= 0.6:
            fill = FILL_STRONG
        else:
            fill = FILL_WEAK
        _fill_row(ws, r, start_col, fill)

        dv_values = list(dict.fromkeys(labels + ["（手动指定）"]))
        _attach_dv(ws, r, start_col, dv_values)

    wb.save(out_path)


def _fill_row(ws, r: int, start_col: int, fill: PatternFill) -> None:
    for i in range(3):
        ws.cell(row=r, column=start_col + i).fill = fill


def _attach_dv(ws, r: int, start_col: int, values: list[str]) -> None:
    dv = DataValidation(
        type="list",
        formula1=_dv_formula(values),
        allow_blank=True,
        showDropDown=False,
    )
    dv.add(ws.cell(row=r, column=start_col).coordinate)
    ws.add_data_validation(dv)


def _dv_formula(values: list[str]) -> str:
    """Data validation list 的 formula1 字符串。Excel 用英文逗号分隔，
    单元内引号转义。单元格值不能含英文逗号（会被当分隔），若含则整体
    失败——本工具的候选字符串不会含英文逗号。"""
    safe = [v.replace('"', '""').replace(",", "，") for v in values]
    return '"' + ",".join(safe) + '"'


# ------------- 主流程 ------------- #

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input", help="空白设计书 .xls/.xlsx 路径")
    ap.add_argument("--project", required=True, help="使用哪个项目的知识库")
    ap.add_argument("--schema", help="输入 schema yaml 路径（默认 <stem>.schema.yaml）")
    ap.add_argument("--out", help="输出 .xlsx 路径（默认 <stem>_候选.xlsx）")
    args = ap.parse_args()

    in_path = Path(args.input).resolve()
    if not in_path.exists():
        print(f"not found: {in_path}", file=sys.stderr)
        return 1
    schema_path = Path(args.schema).resolve() if args.schema else None

    kb_path = ROOT / "projects" / args.project / "knowledge" / "ifs.db"
    if not kb_path.exists():
        print(f"knowledge base not found: {kb_path}", file=sys.stderr)
        return 1

    schema = load_input_schema(in_path, schema_path)
    xlsx_path = ensure_xlsx(in_path)
    fields, if_meta = read_blank_book(xlsx_path, schema)

    out_path = (
        Path(args.out).resolve()
        if args.out
        else in_path.with_name(in_path.stem + "_候选.xlsx")
    )
    shutil.copy(xlsx_path, out_path)

    project_dir = ROOT / "projects" / args.project
    business_dict = load_business_dict(project_dir)

    kb = sqlite3.connect(kb_path)
    kb.row_factory = sqlite3.Row
    norm_index = build_norm_index(kb)
    struct_dict = build_struct_field_dict(kb)

    # Pass 1：先对每个字段做直接命中（不含推测），用于构建上下文聚集
    direct_pairs = [
        (f, pass1_candidates(f, kb, norm_index) if not f.get("skip") else [])
        for f in fields
    ]
    ctx = context_structs(direct_pairs)
    ctx_structs = [s for s, _ in ctx.most_common(5)]

    # 为 AI 推测准备本书字段分布摘要
    book_field_summary = _build_book_field_summary(fields, direct_pairs, business_dict)
    enable_ai = os.environ.get("IF_MAP_DISABLE_AI") != "1"
    if enable_ai:
        print(f"AI 推测启用（{sum(1 for _,d in direct_pairs if not d and not _.get('skip'))} 个字段候选 AI 推测，每字段 ≤{AI_TIMEOUT_SEC}s）", file=sys.stderr)

    # Pass 2：对每个字段决定最终 entries（考虑复合 + 推测）
    results: list[dict] = []
    for f, direct in direct_pairs:
        if f.get("skip"):
            results.append({"field": f, "entries": [(None, [], 0.0, None)], "composite": False})
            continue

        # 复合检测（仅在直接命中为空时启用，避免打扰正常字段）
        if not direct:
            sub_fields = split_composite(f)
            if sub_fields:
                sub_entries = []
                for sf in sub_fields:
                    sf_direct = pass1_candidates(sf, kb, norm_index)
                    if sf_direct:
                        cands = sf_direct
                        skip = None
                    else:
                        skip = check_skip_patterns(sf.get("ext_name"), business_dict)
                        cands = [] if skip else pass1_speculate(sf, kb, business_dict, struct_dict, ctx_structs)
                        cands = _diversify_speculate(cands, limit=TOP_N) if cands else []
                    total_w = sum(c["weighted_freq"] for c in cands) or 1.0
                    sub_entries.append((sf.get("_composite_label") or sf.get("ext_name"), cands, total_w, skip))
                results.append({"field": f, "entries": sub_entries, "composite": True})
                continue

        # 非复合：若无直接命中则尝试 AI 推测，失败降规则
        cands, skip_reason = resolve_candidates(
            f, kb, norm_index, business_dict, struct_dict, ctx_structs,
            ctx_counter=ctx,
            book_field_summary=book_field_summary,
            if_meta=if_meta,
            enable_ai=enable_ai,
        )
        # 若有直接命中，cands 就是 direct
        if direct:
            cands = direct
        total_w = sum(c["weighted_freq"] for c in cands) or 1.0
        results.append({
            "field": f,
            "entries": [(None, cands, total_w, skip_reason)],
            "composite": False,
        })

    render_excel(out_path, fields, results, ctx, schema, out_path, kb)

    n_direct = sum(
        1 for item in results
        if not item["composite"] and item["entries"][0][1]
        and item["entries"][0][1][0].get("origin") == "history"
    )
    n_spec = sum(
        1 for item in results
        if not item["composite"] and item["entries"][0][1]
        and item["entries"][0][1][0].get("origin") == "speculate"
    )
    n_ai = sum(
        1 for item in results
        if not item["composite"] and item["entries"][0][1]
        and item["entries"][0][1][0].get("origin") == "ai_speculate"
    )
    n_comp = sum(1 for item in results if item["composite"])
    n_skip = sum(1 for f in fields if f.get("skip"))
    n_none = sum(
        1 for item in results
        if not item["composite"] and not item["field"].get("skip")
        and not item["entries"][0][1]
    )
    print(f"wrote → {out_path}")
    print(
        f"fields: {len(fields)} | 直接命中: {n_direct} | AI推测: {n_ai} | 规则推测: {n_spec} | "
        f"复合: {n_comp} | 无候选: {n_none} | 跳过: {n_skip}"
    )
    print(f"ctx top: {dict(ctx.most_common(5))}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
