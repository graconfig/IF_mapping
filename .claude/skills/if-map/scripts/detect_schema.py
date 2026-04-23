"""从一份待转换的空白设计书样本探测表格结构，生成 <stem>.schema.yaml 草稿。

要找的：主工作表、header_row、单边字段块（外部系统字段）、可选 remark 列。

用法：
  python3 scripts/detect_schema.py <blank.xls> [--if-name "业务名"] [--out schema.yaml]
"""
from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
import unicodedata
import warnings
from pathlib import Path
from typing import Any

import openpyxl
import yaml

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ---- 语义同义词集 ----
FIELD_SEMANTICS_VOCAB: dict[str, list[str]] = {
    "no":     ["№", "No.", "No", "NO", "序号", "项次", "#"],
    "name":   ["項目名称", "項目名", "Field Name", "Column Name", "字段名称", "字段名", "名称"],
    "struct": ["構造", "Table", "表名", "テーブル", "构造", "Structure"],
    "tech":   ["技術名称", "技術名", "項目コード", "Technical Name", "Field Code", "技术名称", "技术名", "字段代码"],
    "length": ["文字数", "桁数", "バイト数", "Leng", "Length", "Len", "Byte", "长度", "バイト"],
    "type":   ["属性", "Type", "データ型", "类型", "型"],
}

# 空白设计书只识别 remark 列（其他辅助列是参考文档 SAP 侧的，blank 不需要）
REMARK_VOCAB = ["備考", "Remark", "备注"]


# ---- 基础工具 ----

def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _match_vocab(text: str | None, vocab: list[str]) -> bool:
    if not text:
        return False
    t = unicodedata.normalize("NFKC", str(text)).strip()
    return any(t == w or t.startswith(w) for w in vocab)


def _semantic_for(text: str | None) -> str | None:
    if not text:
        return None
    for sem, vocab in FIELD_SEMANTICS_VOCAB.items():
        if _match_vocab(text, vocab):
            return sem
    return None


def _ensure_xlsx(path: Path) -> Path:
    if path.suffix.lower() == ".xlsx":
        return path
    tmp = Path(tempfile.mkdtemp(prefix="detect_"))
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(path)],
        capture_output=True, text=True, check=True,
    )
    out = list(tmp.glob("*.xlsx"))
    if not out:
        raise RuntimeError(f"xls2xlsx failed for {path}")
    return out[0]


# ---- 核心探测 ----

def _scan_header_row(ws, max_scan_rows: int = 12) -> tuple[int, list[tuple[int, int, str]]] | None:
    """扫前 max_scan_rows 行，返回语义命中最多的行 + 该行命中列列表。"""
    best: tuple[int, list[tuple[int, int, str]]] | None = None
    best_count = 0
    max_col = min(ws.max_column, 40)
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        hits: list[tuple[int, int, str]] = []
        for c in range(1, max_col + 1):
            sem = _semantic_for(ws.cell(row=r, column=c).value)
            if sem:
                hits.append((r, c, sem))
        unique_sems = {h[2] for h in hits}
        if len(unique_sems) >= 4 and len(hits) > best_count:
            best = (r, hits)
            best_count = len(hits)
    return best


def _find_field_blocks(hits: list[tuple[int, int, str]]) -> list[list[tuple[int, str]]]:
    """根据命中列表，找连续的字段块（同行里相邻列语义序列，至少 3 个不同语义）。"""
    cols_sorted = sorted(hits, key=lambda h: h[1])
    blocks: list[list[tuple[int, str]]] = []
    current: list[tuple[int, str]] = []
    prev_col: int | None = None
    for _, col, sem in cols_sorted:
        if prev_col is None or col - prev_col <= 2:
            current.append((col, sem))
        else:
            if len({s for _, s in current}) >= 3:
                blocks.append(current)
            current = [(col, sem)]
        prev_col = col
    if current and len({s for _, s in current}) >= 3:
        blocks.append(current)
    return blocks


def _block_columns(block: list[tuple[int, str]]) -> dict[str, int]:
    result: dict[str, int] = {}
    for col, sem in block:
        result.setdefault(sem, col)
    return result


def _detect_data_start_row(ws, header_row: int) -> int:
    for r in range(header_row + 1, min(ws.max_row, header_row + 5) + 1):
        if any(_clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 20) + 1)):
            return r
    return header_row + 1


def _find_remark_col(ws, header_row: int) -> int | None:
    """在表头行找 备考/Remark 列。"""
    for c in range(1, ws.max_column + 1):
        t = _clean(ws.cell(row=header_row, column=c).value)
        if t and any(w in str(t) for w in REMARK_VOCAB):
            return c
    return None


# ---- 探测主入口 ----

def detect_blank(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    best: tuple[str, int, list[list[tuple[int, str]]]] | None = None
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        scan = _scan_header_row(ws)
        if not scan:
            continue
        hr, hits = scan
        blocks = _find_field_blocks(hits)
        if not blocks:
            continue
        top_block = blocks[0]
        score = len({s for _, s in top_block}) * 10 + len(top_block)
        if score > best_score:
            best = (name, hr, blocks)
            best_score = score
    if not best:
        raise RuntimeError(f"no blank-book-like sheet found in {xlsx_path.name}")
    sheet_name, header_row, blocks = best
    ws = wb[sheet_name]
    cols = _block_columns(blocks[0])
    remark_col = _find_remark_col(ws, header_row)
    if remark_col:
        cols["remark"] = remark_col

    return {
        "sheet": sheet_name,
        "header_row": header_row,
        "data_start_row": _detect_data_start_row(ws, header_row),
        "columns": cols,
    }


# ---- yaml 草稿输出 ----

def _col_letter(idx: int) -> str:
    s = ""
    n = idx
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _dump_yaml(data: dict, header_comment: str = "") -> str:
    body = yaml.safe_dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False)
    return (header_comment + "\n" if header_comment else "") + body


def render_blank_yaml(detected: dict, if_name_hint: str = "") -> str:
    # 探测器内部语义 → fill_book 期待的 schema key
    SEM_TO_SCHEMA_KEY = {
        "no": "ext_no", "name": "ext_name", "tech": "ext_tech",
        "type": "ext_type", "length": "ext_len", "remark": "remark",
    }
    cols = {SEM_TO_SCHEMA_KEY.get(sem, sem): _col_letter(c)
            for sem, c in detected["columns"].items()}
    data = {
        "sheet": detected["sheet"],
        "header_row": detected["header_row"],
        "data_start_row": detected["data_start_row"],
        "columns": cols,
        "skip": {"names": ["予備"], "techs": ["FIL1", "FIL2", "FIL3"]},
        "if_meta": {
            "ifid_guess": "",
            "if_name": if_name_hint,
            "direction": "external_to_sap",
            "counterpart_hint": None,
        },
    }
    header = (
        "# 自动探测生成 — 请人工复核\n"
        f"# 样本工作表：{detected['sheet']}\n"
        f"# 检出字段语义列：{detected['columns']}\n"
    )
    return _dump_yaml(data, header)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="空白设计书 Excel 路径（.xls / .xlsx 都行）")
    ap.add_argument("--if-name", default="", help="IF 业务名提示")
    ap.add_argument("--out", help="yaml 输出路径（默认打印到 stdout）")
    args = ap.parse_args()

    src = Path(args.xlsx).resolve()
    if not src.exists():
        print(f"not found: {src}", file=sys.stderr)
        return 1
    xlsx = _ensure_xlsx(src)

    detected = detect_blank(xlsx)
    yaml_text = render_blank_yaml(detected, args.if_name)

    if args.out:
        Path(args.out).write_text(yaml_text, encoding="utf-8")
        print(f"wrote → {args.out}")
    else:
        print(yaml_text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
