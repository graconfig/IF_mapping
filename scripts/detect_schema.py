"""从一份 Excel 样本探测表格结构，生成 config.yaml 或 schema.yaml 草稿。

两种模式：
  --as reference  : 为历史设计书（参考文档）生成 projects/<name>/config.yaml 草稿
                    要找：映射表工作表、header_row、两个字段块（源+目标）、SAP 侧 label
  --as blank      : 为待转换的空白设计书生成同目录下 <stem>.schema.yaml 草稿
                    要找：主工作表、header_row、单边字段块

用法：
  python3 scripts/detect_schema.py --as reference <sample.xlsx> [--out config.yaml]
  python3 scripts/detect_schema.py --as blank <blank.xls> [--out schema.yaml]
"""
from __future__ import annotations

import argparse
import re
import subprocess
import sys
import tempfile
import unicodedata
import warnings
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
import yaml

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---- 语义同义词集 ----
# 每个语义提供多语言同义词；同义词按优先级排列（供 config 写草稿时选首选）
FIELD_SEMANTICS_VOCAB: dict[str, list[str]] = {
    "no":     ["№", "No.", "No", "NO", "序号", "项次", "#"],
    "name":   ["項目名称", "項目名", "Field Name", "Column Name", "字段名称", "字段名", "名称"],
    "struct": ["構造", "Table", "表名", "テーブル", "构造", "Structure"],
    "tech":   ["技術名称", "技術名", "項目コード", "Technical Name", "Field Code", "技术名称", "技术名", "字段代码"],
    "length": ["文字数", "桁数", "バイト数", "Leng", "Length", "Len", "Byte", "长度", "バイト"],
    "type":   ["属性", "Type", "データ型", "类型", "型"],
}

# 映射表工作表名关键词
MAPPING_SHEET_KEYWORDS = ["項目マッピング", "マッピング", "Mapping", "mapping", "映射", "Map"]

# SAP 侧 label 正则（同 config.yaml 默认 target_side）
SAP_LABEL_PATTERNS = [r"部品\s*SAP", r"ＳＡＰ"]

# 辅助列同义词（参考文档）
AUX_VOCAB = {
    "conv_spec":        ["変換仕様"],
    "conv_current":     ["現行編集仕様"],
    "sap_digits":       ["桁数"],
    "sap_code_system":  ["コード体系"],
    "sap_supplement":   ["補足・その他", "補足"],
    "unrealizable_no":  ["No."],
    "unrealizable_class": ["分類"],
    "remark":           ["備考", "Remark", "备注"],
}


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _match_vocab(text: str | None, vocab: list[str]) -> bool:
    if not text:
        return False
    t = unicodedata.normalize("NFKC", str(text)).strip()
    # 精确或 startswith 命中
    return any(t == w or t.startswith(w) for w in vocab)


def _semantic_for(text: str | None) -> str | None:
    if not text:
        return None
    for sem, vocab in FIELD_SEMANTICS_VOCAB.items():
        if _match_vocab(text, vocab):
            return sem
    return None


# ---- .xls 自动转 .xlsx ----

def _ensure_xlsx(path: Path) -> Path:
    if path.suffix.lower() == ".xlsx":
        return path
    tmp = Path(tempfile.mkdtemp(prefix="detect_"))
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(path)],
        capture_output=True, text=True, check=True,
    )
    out = list(tmp.glob("*.xlsx"))
    if not out:
        raise RuntimeError(f"xls2xlsx failed for {path}")
    return out[0]


# ---- 核心探测：找 header_row + 字段块 ----

def _scan_header_row(ws, max_scan_rows: int = 12) -> tuple[int, list[tuple[int, int, str]]] | None:
    """扫描前 max_scan_rows 行，返回语义命中最多的行 + 该行命中列列表。

    返回 (row_idx, [(col, semantic), ...]) 或 None（未找到）。
    """
    best: tuple[int, list[tuple[int, int, str]]] | None = None
    best_count = 0
    max_col = min(ws.max_column, 40)
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        hits: list[tuple[int, int, str]] = []
        for c in range(1, max_col + 1):
            sem = _semantic_for(ws.cell(row=r, column=c).value)
            if sem:
                hits.append((r, c, sem))
        # 至少命中 4 个语义才像是表头
        unique_sems = {h[2] for h in hits}
        if len(unique_sems) >= 4 and len(hits) > best_count:
            best = (r, hits)
            best_count = len(hits)
    return best


def _find_field_blocks(hits: list[tuple[int, int, str]]) -> list[list[tuple[int, str]]]:
    """根据命中列表，找连续的字段块（同一行里相邻列的语义序列）。

    一个"块"：相邻列（允许 1 列缺口）命中至少 3 个不同语义。
    """
    cols_sorted = sorted(hits, key=lambda h: h[1])
    blocks: list[list[tuple[int, str]]] = []
    current: list[tuple[int, str]] = []
    prev_col: int | None = None
    for _, col, sem in cols_sorted:
        if prev_col is None or col - prev_col <= 2:
            current.append((col, sem))
        else:
            if len({s for _, s in current}) >= 3:
                blocks.append(current)
            current = [(col, sem)]
        prev_col = col
    if current and len({s for _, s in current}) >= 3:
        blocks.append(current)
    return blocks


def _block_columns(block: list[tuple[int, str]]) -> dict[str, int]:
    """字段块 → {语义: 列号}。同语义重复时取第一个。"""
    result: dict[str, int] = {}
    for col, sem in block:
        result.setdefault(sem, col)
    return result


def _detect_data_start_row(ws, header_row: int) -> int:
    """表头下第一行非空即数据起始。"""
    for r in range(header_row + 1, min(ws.max_row, header_row + 5) + 1):
        if any(_clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 20) + 1)):
            return r
    return header_row + 1


def _find_sap_side(ws, header_row: int, blocks: list[list[tuple[int, str]]]) -> int | None:
    """在 header_row-1 行找哪一个 block 的起点所在列区域命中 SAP label。
    返回命中 block 的索引（0 或 1）；找不到返回 None。
    """
    label_row = header_row - 1
    if label_row < 1:
        return None
    patterns = [re.compile(p, re.IGNORECASE) for p in SAP_LABEL_PATTERNS]
    # A 列（col=1）的 label 覆盖左侧整块
    left_label = _clean(ws.cell(row=label_row, column=1).value) or ""
    if any(p.search(left_label) for p in patterns):
        return 0
    for i, b in enumerate(blocks):
        start_col = b[0][0]
        lab = _clean(ws.cell(row=label_row, column=start_col).value) or ""
        if any(p.search(lab) for p in patterns):
            return i
    return None


def _find_aux_columns(ws, header_row: int, label_row: int | None = None) -> dict[str, int | None]:
    """扫 header_row 与 label_row 的单元格文本，定位辅助列。"""
    label_row = label_row or header_row - 1
    result: dict[str, int | None] = {k: None for k in AUX_VOCAB}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        t5 = _clean(ws.cell(row=max(1, label_row), column=c).value)
        t6 = _clean(ws.cell(row=header_row, column=c).value)
        joined = " ".join(str(x) for x in (t5, t6) if x)
        for key, vocab in AUX_VOCAB.items():
            if result[key] is None:
                if t6 and str(t6).strip() in vocab:
                    result[key] = c
                elif any(v in joined for v in vocab):
                    result[key] = c
    return result


def _is_mapping_sheet_name(name: str) -> bool:
    return any(k in name for k in MAPPING_SHEET_KEYWORDS)


# ---- mode: reference ----

def detect_reference(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    candidates: list[tuple[str, int, list[list[tuple[int, str]]]]] = []
    for name in wb.sheetnames:
        if not _is_mapping_sheet_name(name):
            continue
        ws = wb[name]
        scan = _scan_header_row(ws)
        if not scan:
            continue
        hr, hits = scan
        blocks = _find_field_blocks(hits)
        if len(blocks) >= 2:
            candidates.append((name, hr, blocks))
    if not candidates:
        raise RuntimeError(f"no valid mapping sheet found in {xlsx_path.name}")

    # 取命中质量最高的（block 数最多 + 语义覆盖最好）
    candidates.sort(key=lambda x: -(len(x[2]) + sum(len({s for _, s in b}) for b in x[2])))
    sheet_name, header_row, blocks = candidates[0]
    ws = wb[sheet_name]

    left_block, right_block = blocks[0], blocks[1]
    sap_idx = _find_sap_side(ws, header_row, [left_block, right_block])
    sap_side = {0: "left", 1: "right", None: "right"}[sap_idx]

    data_start = _detect_data_start_row(ws, header_row)
    aux = _find_aux_columns(ws, header_row)

    return {
        "sheet_name_sample": sheet_name,
        "mapping_sheet_name_regex": _regex_from_names([s for s in wb.sheetnames if _is_mapping_sheet_name(s)]),
        "header_row": header_row,
        "data_start_row": data_start,
        "label_row": max(1, header_row - 1),
        "left_block": _block_columns(left_block),
        "right_block": _block_columns(right_block),
        "sap_side": sap_side,
        "aux_columns": {k: v for k, v in aux.items() if v},
    }


def _regex_from_names(names: list[str]) -> str:
    """根据一组实际工作表名，推出一个安全的匹配正则。

    策略：找出 MAPPING_SHEET_KEYWORDS 里命中所有 names 的最长关键词。
    这样无论样本是 "項目マッピング(受信)" 还是 "項目マッピング (01)"，
    都能统一命中为 "項目マッピング"，不会把括号后缀 lock 死。
    """
    if not names:
        return "項目マッピング"
    for kw in sorted(MAPPING_SHEET_KEYWORDS, key=len, reverse=True):
        if all(kw in n for n in names):
            return re.escape(kw)
    return "マッピング"


# ---- mode: blank ----

def detect_blank(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    best: tuple[str, int, list[tuple[int, int, str]], list[list[tuple[int, str]]]] | None = None
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        scan = _scan_header_row(ws)
        if not scan:
            continue
        hr, hits = scan
        blocks = _find_field_blocks(hits)
        if not blocks:
            continue
        top_block = blocks[0]
        score = len({s for _, s in top_block}) * 10 + len(top_block)
        if score > best_score:
            best = (name, hr, hits, blocks)
            best_score = score
    if not best:
        raise RuntimeError(f"no blank-book-like sheet found in {xlsx_path.name}")
    sheet_name, header_row, _hits, blocks = best
    ws = wb[sheet_name]
    cols = _block_columns(blocks[0])
    # 补充辅助列（备注）
    aux = _find_aux_columns(ws, header_row)
    if aux.get("remark"):
        cols["remark"] = aux["remark"]
    data_start = _detect_data_start_row(ws, header_row)

    return {
        "sheet": sheet_name,
        "header_row": header_row,
        "data_start_row": data_start,
        "columns": cols,
    }


# ---- yaml 草稿输出 ----

def _col_letter(idx: int) -> str:
    s = ""
    n = idx
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _dump_yaml(data: dict, header_comment: str = "") -> str:
    body = yaml.safe_dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False)
    return (header_comment + "\n" if header_comment else "") + body


def render_reference_yaml(detected: dict, project_name: str = "unknown") -> str:
    data = {
        "name": project_name,
        "description": "自动生成，请人工复核",
        "sources_dir": "sources",
        "out_dir": "knowledge",
        "mapping_sheet": {
            "name_regex": detected["mapping_sheet_name_regex"],
            "header_row": detected["header_row"],
            "data_start_row": detected["data_start_row"],
            "label_row": detected["label_row"],
        },
        "field_semantics": FIELD_SEMANTICS_VOCAB,
        "target_side": {
            "label_patterns": [r"部品\s*SAP", "ＳＡＰ"],
        },
        "aux_columns": AUX_VOCAB,
        "direction_rules": {
            "sheet_suffix": {"(受信)": "external_to_sap", "(返信)": "sap_to_external"},
            "file_marker":  {"※受信": "external_to_sap", "※送信": "sap_to_external"},
            "fallback_by_sap_side": True,
        },
        "if_meta": {
            "ifid_cell":    [2, 1],
            "if_name_cell": [2, 3],
        },
    }
    header = (
        "# 自动探测生成 — 请人工复核\n"
        f"# 样本工作表：{detected['sheet_name_sample']}\n"
        f"# 检出字段块：left={detected['left_block']}\n"
        f"#           right={detected['right_block']}\n"
        f"# 推断 SAP 侧：{detected['sap_side']}\n"
        f"# [TODO] 请确认 if_meta.ifid_cell 与 if_name_cell 所在单元格\n"
    )
    return _dump_yaml(data, header)


def render_blank_yaml(detected: dict, if_name_hint: str = "") -> str:
    cols = {sem: _col_letter(c) for sem, c in detected["columns"].items()}
    data = {
        "sheet": detected["sheet"],
        "header_row": detected["header_row"],
        "data_start_row": detected["data_start_row"],
        "columns": cols,
        "skip": {
            "names": ["予備"],
            "techs": ["FIL1", "FIL2", "FIL3"],
        },
        "if_meta": {
            "ifid_guess": "",
            "if_name": if_name_hint,
            "direction": "external_to_sap",
            "counterpart_hint": None,
        },
    }
    header = (
        "# 自动探测生成 — 请人工复核\n"
        f"# 样本工作表：{detected['sheet']}\n"
        f"# 检出字段语义列：{detected['columns']}\n"
    )
    return _dump_yaml(data, header)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="样本 Excel 路径")
    ap.add_argument("--as", dest="mode", choices=["reference", "blank"], required=True,
                    help="reference=参考文档（历史映射书）；blank=待转换空白设计书")
    ap.add_argument("--out", help="yaml 草稿输出路径（默认打印到 stdout）")
    ap.add_argument("--project", default="unknown", help="项目名（reference 模式用）")
    ap.add_argument("--if-name", default="", help="IF 名提示（blank 模式用）")
    args = ap.parse_args()

    src = Path(args.xlsx).resolve()
    if not src.exists():
        print(f"not found: {src}", file=sys.stderr)
        return 1
    xlsx = _ensure_xlsx(src)

    if args.mode == "reference":
        detected = detect_reference(xlsx)
        yaml_text = render_reference_yaml(detected, args.project)
    else:
        detected = detect_blank(xlsx)
        yaml_text = render_blank_yaml(detected, args.if_name)

    if args.out:
        Path(args.out).write_text(yaml_text, encoding="utf-8")
        print(f"wrote → {args.out}")
    else:
        print(yaml_text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
