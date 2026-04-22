"""按项目配置 (config.yaml) 从一份接口设计书 Excel 抽取字段映射记录。

用法：
  python3 scripts/extract_mapping.py --config projects/<name>/config.yaml <xlsx 路径> [--out <jsonl>]

输出字段统一使用 external_* / sap_* —— 抽取器根据 config.target_side 判定哪一侧是 SAP。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import yaml

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

SEMANTIC_ORDER = ("no", "name", "struct", "tech", "length", "type")


def load_config(path: str | Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def _match_semantic(cell_text: str | None, vocab: list[str]) -> bool:
    if not cell_text:
        return False
    t = str(cell_text).strip()
    return any(t == w for w in vocab)


def _semantic_for_cell(cell_text: str | None, field_semantics: dict[str, list[str]]) -> str | None:
    """返回该 cell 文本命中的语义 key；不命中则 None。"""
    if not cell_text:
        return None
    t = str(cell_text).strip()
    for sem, vocab in field_semantics.items():
        if t in vocab:
            return sem
    return None


def _find_field_blocks(ws, config: dict) -> tuple[tuple[int, ...], tuple[int, ...]] | None:
    """扫描 header_row，找两个连续 6 列的语义块 (no,name,struct,tech,length,type)。"""
    header_row = config["mapping_sheet"]["header_row"]
    semantics = config["field_semantics"]
    max_col = ws.max_column
    # 遍历每个起点，尝试匹配完整序列
    cell_sems = [
        _semantic_for_cell(ws.cell(row=header_row, column=c).value, semantics)
        for c in range(1, max_col + 1)
    ]
    blocks: list[tuple[int, ...]] = []
    c = 0
    while c < max_col - 5:
        # cell_sems[c] 对应列 c+1
        if cell_sems[c] == SEMANTIC_ORDER[0]:  # "no"
            seq = cell_sems[c:c + 6]
            if tuple(seq) == SEMANTIC_ORDER:
                blocks.append(tuple(range(c + 1, c + 7)))
                c += 6
                continue
        c += 1
    if len(blocks) >= 2:
        return blocks[0], blocks[1]
    return None


def _find_aux_cols(ws, config: dict) -> dict[str, int | None]:
    """按 aux_columns 配置，在 row=label_row 和 header_row 扫描关键字，找列号。"""
    label_row = config["mapping_sheet"]["label_row"]
    header_row = config["mapping_sheet"]["header_row"]
    aux_cfg = config.get("aux_columns", {})
    result: dict[str, int | None] = {k: None for k in aux_cfg}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        t5 = _clean(ws.cell(row=label_row, column=c).value)
        t6 = _clean(ws.cell(row=header_row, column=c).value)
        joined = " ".join(str(x) for x in (t5, t6) if x)
        for key, vocab in aux_cfg.items():
            if result[key] is not None:
                continue
            # 精确或包含匹配：row6 精确，或联合文本包含
            if t6 and str(t6).strip() in vocab:
                result[key] = c
            elif any(v in joined for v in vocab):
                result[key] = c
    return result


def _detect_sap_side(
    ws, left_cols: tuple[int, ...], right_cols: tuple[int, ...], config: dict
) -> tuple[str, str]:
    """判定 SAP 在哪一侧。返回 (sap_side, counterpart_name)。sap_side ∈ {"left","right"}。"""
    label_row = config["mapping_sheet"]["label_row"]
    patterns = [re.compile(p, re.IGNORECASE) for p in config["target_side"]["label_patterns"]]

    # 左侧 label: col=1；右侧 label: col=right_cols[0]
    left_label = _clean(ws.cell(row=label_row, column=1).value) or ""
    right_label = _clean(ws.cell(row=label_row, column=right_cols[0]).value) or ""

    left_hit = any(p.search(left_label) for p in patterns)
    right_hit = any(p.search(right_label) for p in patterns)

    if left_hit and not right_hit:
        return "left", right_label or "(unknown)"
    if right_hit and not left_hit:
        return "right", left_label or "(unknown)"
    # 两侧都命中或都不命中：启发式降级，默认右为 SAP
    if right_hit:
        return "right", left_label or "(unknown)"
    if left_hit:
        return "left", right_label or "(unknown)"
    return "right", left_label or right_label or "(unknown)"


def _infer_direction(sheet_name: str, file_name: str, sap_side: str, config: dict) -> str:
    rules = config.get("direction_rules", {})
    for marker, direction in (rules.get("sheet_suffix") or {}).items():
        if marker in sheet_name:
            return direction
    for marker, direction in (rules.get("file_marker") or {}).items():
        if marker in file_name:
            return direction
    if rules.get("fallback_by_sap_side", True):
        if sap_side == "left":
            return "sap_to_external"
        if sap_side == "right":
            return "external_to_sap"
    return "unknown"


def _pack_side(ws, r: int, cols: tuple[int, ...]) -> dict[str, Any]:
    keys = ("no", "name", "struct", "tech", "len", "attr")
    return {k: _clean(ws.cell(row=r, column=c).value) for k, c in zip(keys, cols)}


def _row_is_empty(ws, r: int, watched_cols: list[int]) -> bool:
    return all(_clean(ws.cell(row=r, column=c).value) is None for c in watched_cols)


def extract_sheet(ws, source_file: str, config: dict) -> list[dict[str, Any]]:
    blocks = _find_field_blocks(ws, config)
    if blocks is None:
        return []
    left_cols, right_cols = blocks
    aux = _find_aux_cols(ws, config)
    sap_side, counterpart = _detect_sap_side(ws, left_cols, right_cols, config)

    label_row = config["mapping_sheet"]["label_row"]
    ifid_rc = config["if_meta"]["ifid_cell"]
    name_rc = config["if_meta"]["if_name_cell"]
    ifid = _clean(ws.cell(row=ifid_rc[0], column=ifid_rc[1]).value)
    if_name = _clean(ws.cell(row=name_rc[0], column=name_rc[1]).value)
    sap_label = _clean(
        ws.cell(row=label_row, column=right_cols[0] if sap_side == "right" else 1).value
    )
    direction = _infer_direction(ws.title, source_file, sap_side, config)

    sap_cols = right_cols if sap_side == "right" else left_cols
    ext_cols = left_cols if sap_side == "right" else right_cols

    watched = list(left_cols) + list(right_cols) + [v for v in aux.values() if v]
    data_start = config["mapping_sheet"]["data_start_row"]

    records: list[dict[str, Any]] = []
    for r in range(data_start, ws.max_row + 1):
        if _row_is_empty(ws, r, watched):
            continue
        ext = _pack_side(ws, r, ext_cols)
        sap = _pack_side(ws, r, sap_cols)

        def _aux(key: str):
            col = aux.get(key)
            return _clean(ws.cell(row=r, column=col).value) if col else None

        rec = {
            "source_file": source_file,
            "sheet": ws.title,
            "row_idx": r,
            "ifid": ifid,
            "if_name": if_name,
            "counterpart_system": counterpart,
            "sap_side_label": sap_label,
            "sap_side": sap_side,
            "direction": direction,
            "ext_no": ext["no"], "ext_name": ext["name"], "ext_struct": ext["struct"],
            "ext_tech": ext["tech"], "ext_len": ext["len"], "ext_attr": ext["attr"],
            "sap_no": sap["no"], "sap_name": sap["name"], "sap_struct": sap["struct"],
            "sap_tech": sap["tech"], "sap_len": sap["len"], "sap_attr": sap["attr"],
            "conv_spec":    _aux("conv_spec"),
            "conv_current": _aux("conv_current"),
            "sap_digits":   _aux("sap_digits"),
            "sap_code_system": _aux("sap_code_system"),
            "sap_supplement":  _aux("sap_supplement"),
            "unrealizable_no":    _aux("unrealizable_no"),
            "unrealizable_class": _aux("unrealizable_class"),
            "remark":       _aux("remark"),
        }
        records.append(rec)
    return records


def extract_file(path: Path, config: dict) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_re = re.compile(config["mapping_sheet"]["name_regex"])
    records: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        if not sheet_re.search(name):
            continue
        records.extend(extract_sheet(wb[name], path.name, config))
    return records


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="设计书 Excel 路径")
    ap.add_argument("--config", required=True, help="项目配置 yaml 路径")
    ap.add_argument("--out", help="输出 jsonl 路径（默认 projects/<name>/knowledge/pilot/<stem>.jsonl）")
    args = ap.parse_args(argv[1:])

    src = Path(args.xlsx)
    if not src.exists():
        print(f"not found: {src}", file=sys.stderr)
        return 1
    config = load_config(args.config)
    records = extract_file(src, config)

    if args.out:
        out_path = Path(args.out)
    else:
        cfg_root = Path(args.config).resolve().parent
        out_path = cfg_root / config.get("out_dir", "knowledge") / "pilot" / (
            re.sub(r"[^A-Za-z0-9_.-]", "_", src.stem) + ".jsonl"
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"wrote {len(records)} records → {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
